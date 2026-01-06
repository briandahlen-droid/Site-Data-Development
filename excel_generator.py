"""
Excel Report Generator for Florida Parcel Lookup
Generates formatted Excel reports matching the reference format
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def generate_parcel_report(property_data, zoning_data=None, output_path="parcel_report.xlsx", sections=None):
    """
    Generate formatted Excel report with property and zoning data.
    
    Args:
        property_data (dict): Property information from API lookup
        zoning_data (dict): Optional zoning requirements from Municode
        output_path (str): Path for output Excel file
        sections (dict): Dict of section flags (which sections to include)
    
    Returns:
        str: Path to generated file
    """
    
    # Default: include all sections if not specified
    if sections is None:
        sections = {
            'property_info': True,
            'site_characteristics': True,
            'zoning_land_use': True,
            'building_requirements': True,
            'parking_requirements': True,
            'assessment_values': True,
            'sales_history': True,
            'links_references': True
        }
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Parcel Information"
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
    
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(name='Arial', size=11, bold=True)
    
    label_font = Font(name='Arial', size=10, bold=True)
    value_font = Font(name='Arial', size=10)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    alignment_center = Alignment(horizontal='center', vertical='center')
    
    # Current row tracker
    current_row = 1
    
    # Helper function to add a section header
    def add_section_header(title):
        nonlocal current_row
        ws.merge_cells(f'A{current_row}:B{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = title
        cell.fill = section_fill
        cell.font = section_font
        cell.alignment = alignment_center
        cell.border = border
        current_row += 1
    
    # Helper function to add a data row
    def add_data_row(label, value):
        nonlocal current_row
        label_cell = ws[f'A{current_row}']
        value_cell = ws[f'B{current_row}']
        
        label_cell.value = label
        label_cell.font = label_font
        label_cell.alignment = alignment_left
        label_cell.border = border
        
        value_cell.value = value if value else "Not Available"
        value_cell.font = value_font
        value_cell.alignment = alignment_left
        value_cell.border = border
        
        current_row += 1
    
    # Main Header
    ws.merge_cells(f'A1:B1')
    main_header = ws['A1']
    main_header.value = "PROPERTY INFORMATION REPORT"
    main_header.fill = header_fill
    main_header.font = header_font
    main_header.alignment = alignment_center
    main_header.border = border
    current_row = 2
    
    # Empty row for spacing
    current_row += 1
    
    # PROPERTY INFORMATION SECTION
    if sections.get('property_info', True):
        add_section_header("PROPERTY INFORMATION")
        
        add_data_row("Owner Name", property_data.get('owner', ''))
        add_data_row("Owner Address", property_data.get('owner_address', ''))
        
        # Construct full owner city/state/zip
        owner_location = ""
        if property_data.get('owner_city'):
            owner_location = f"{property_data.get('owner_city', '')}, {property_data.get('owner_state', '')} {property_data.get('owner_zip', '')}"
        add_data_row("Owner City/State/Zip", owner_location)
        
        add_data_row("Property Address", property_data.get('address', ''))
        
        # Construct full property city/zip
        property_location = f"{property_data.get('city', '')}, FL {property_data.get('zip', '')}"
        add_data_row("Property City/Zip", property_location)
        
        add_data_row("Legal Description", property_data.get('legal_description', ''))
        if property_data.get('legal_description2'):
            add_data_row("Legal Description (cont.)", property_data.get('legal_description2', ''))
        
        # Empty row for spacing
        current_row += 1
    
    # SITE CHARACTERISTICS SECTION
    if sections.get('site_characteristics', True):
        add_section_header("SITE CHARACTERISTICS")
    
    acres = property_data.get('acres', 0)
    add_data_row("Site Area", f"{acres:.2f} acres" if acres else "Not Available")
    
    sqft = property_data.get('area_sqft', 0)
    if sqft:
        add_data_row("Site Area (sq ft)", f"{sqft:,.0f} sq ft")
    
    add_data_row("Section/Township/Range", 
                 f"S{property_data.get('section', '')}-T{property_data.get('township', '')}-R{property_data.get('range', '')}")
    
    if property_data.get('subdivision'):
        add_data_row("Subdivision", property_data.get('subdivision', ''))
    
    if property_data.get('block') or property_data.get('lot'):
        add_data_row("Block/Lot", f"Block {property_data.get('block', 'N/A')} / Lot {property_data.get('lot', 'N/A')}")
    
    # Empty row for spacing
    current_row += 1
    
    # ZONING AND LAND USE SECTION
    add_section_header("ZONING & LAND USE")
    
    add_data_row("Current Zoning", property_data.get('zoning', 'Contact jurisdiction'))
    add_data_row("Land Use Description", property_data.get('land_use', ''))
    add_data_row("Land Use Code", property_data.get('land_use_code', ''))
    
    # If zoning data is provided from Municode
    if zoning_data:
        add_data_row("Future Land Use", zoning_data.get('future_land_use', 'TBD'))
        add_data_row("FEMA Flood Zone", zoning_data.get('fema_flood_zone', 'TBD'))
    else:
        add_data_row("Future Land Use", "Requires separate lookup")
        add_data_row("FEMA Flood Zone", "Requires separate lookup")
    
    # Empty row for spacing
    current_row += 1
    
    # BUILDING REQUIREMENTS SECTION (if zoning data available)
    if zoning_data:
        add_section_header("BUILDING REQUIREMENTS")
        
        setbacks = zoning_data.get('setbacks', {})
        add_data_row("Setback - Front", f"{setbacks.get('front', 'TBD')} ft" if setbacks.get('front') else "TBD")
        add_data_row("Setback - Rear", f"{setbacks.get('rear', 'TBD')} ft" if setbacks.get('rear') else "TBD")
        add_data_row("Setback - Side", f"{setbacks.get('side', 'TBD')} ft" if setbacks.get('side') else "TBD")
        add_data_row("Setback - Street Side", f"{setbacks.get('street_side', 'TBD')} ft" if setbacks.get('street_side') else "TBD")
        
        add_data_row("Maximum Building Height", zoning_data.get('max_height', 'TBD'))
        add_data_row("Maximum Lot Coverage", zoning_data.get('max_coverage', 'TBD'))
        
        # Empty row for spacing
        current_row += 1
        
        # PARKING REQUIREMENTS SECTION
        add_section_header("PARKING REQUIREMENTS")
        
        add_data_row("Standard Parking", zoning_data.get('parking_standard', 'TBD'))
        add_data_row("Bicycle Parking", zoning_data.get('bicycle_parking', 'TBD'))
        add_data_row("Accessible Parking", zoning_data.get('accessible_parking', 'Per ADA/FBC'))
    
    # Empty row for spacing
    current_row += 1
    
    # PROPERTY CHARACTERISTICS SECTION
    add_section_header("PROPERTY CHARACTERISTICS")
    
    if property_data.get('year_built'):
        add_data_row("Year Built", str(property_data.get('year_built', '')))
    
    if property_data.get('num_buildings'):
        add_data_row("Number of Buildings", str(property_data.get('num_buildings', '')))
    
    if property_data.get('num_units'):
        add_data_row("Number of Units", str(property_data.get('num_units', '')))
    
    if property_data.get('total_living_area'):
        add_data_row("Total Living Area", f"{property_data.get('total_living_area', ''):,} sq ft")
    
    # Empty row for spacing
    current_row += 1
    
    # ASSESSMENT VALUES SECTION
    add_section_header("ASSESSMENT VALUES")
    
    assessed_land = property_data.get('assessed_land', 0)
    add_data_row("Assessed Land Value", f"${assessed_land:,}" if assessed_land else "Not Available")
    
    assessed_building = property_data.get('assessed_building', 0)
    add_data_row("Assessed Building Value", f"${assessed_building:,}" if assessed_building else "Not Available")
    
    assessed_total = property_data.get('assessed_total', 0)
    add_data_row("Total Assessed Value", f"${assessed_total:,}" if assessed_total else "Not Available")
    
    market_value = property_data.get('market_value', 0)
    if market_value:
        add_data_row("Market Value", f"${market_value:,}")
    
    # Empty row for spacing
    current_row += 1
    
    # SALES HISTORY SECTION
    if property_data.get('sale_date') or property_data.get('sale_amount'):
        add_section_header("SALES HISTORY")
        
        add_data_row("Most Recent Sale Date", property_data.get('sale_date', 'Not Available'))
        
        sale_amount = property_data.get('sale_amount', 0)
        add_data_row("Most Recent Sale Amount", f"${sale_amount:,}" if sale_amount else "Not Available")
        
        # Empty row for spacing
        current_row += 1
    
    # LINKS AND REFERENCES SECTION
    add_section_header("LINKS & REFERENCES")
    
    if property_data.get('parcel_link'):
        add_data_row("Property Appraiser Link", property_data.get('parcel_link', ''))
    
    add_data_row("Source", "Property data from county property appraiser via SWFWMD ArcGIS service")
    
    if zoning_data:
        add_data_row("Zoning Source", f"Municode - {zoning_data.get('jurisdiction', 'County')} Land Development Code")
    
    # Add disclaimer at bottom
    current_row += 2
    ws.merge_cells(f'A{current_row}:B{current_row}')
    disclaimer = ws[f'A{current_row}']
    disclaimer.value = "DISCLAIMER: This report is for informational purposes only. Property and zoning information should be verified with the appropriate jurisdiction. Kimley-Horn is not responsible for errors or omissions in public data sources."
    disclaimer.font = Font(name='Arial', size=8, italic=True)
    disclaimer.alignment = alignment_left
    
    # Save the workbook
    wb.save(output_path)
    
    return output_path


# Example usage
if __name__ == "__main__":
    # Sample property data structure
    sample_property_data = {
        'owner': 'SAMPLE PROPERTY LLC',
        'owner_address': '123 Main Street',
        'owner_city': 'St. Petersburg',
        'owner_state': 'FL',
        'owner_zip': '33701',
        'address': '456 Development Blvd',
        'city': 'Clearwater',
        'zip': '33755',
        'legal_description': 'LOT 1, BLOCK A, SAMPLE SUBDIVISION, ACCORDING TO THE PLAT THEREOF RECORDED IN PLAT BOOK 100, PAGE 50',
        'acres': 2.5,
        'area_sqft': 108900,
        'zoning': 'RMF-16',
        'land_use': 'Multi-Family Residential',
        'land_use_code': '0100',
        'assessed_land': 450000,
        'assessed_building': 1250000,
        'assessed_total': 1700000,
        'section': '12',
        'township': '29S',
        'range': '16E',
        'year_built': 2018,
        'num_buildings': 1,
        'num_units': 24
    }
    
    sample_zoning_data = {
        'future_land_use': 'Urban Residential',
        'fema_flood_zone': 'Zone X',
        'setbacks': {
            'front': 25,
            'rear': 20,
            'side': 10,
            'street_side': 15
        },
        'max_height': '45 feet / 3 stories',
        'max_coverage': '60%',
        'parking_standard': '1.5 spaces per unit',
        'bicycle_parking': '1 space per 4 units',
        'jurisdiction': 'Pinellas County'
    }
    
    output_file = generate_parcel_report(sample_property_data, sample_zoning_data, "sample_parcel_report.xlsx")
    print(f"Sample report generated: {output_file}")
